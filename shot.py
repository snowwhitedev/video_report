from pathlib import Path
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
#from xlwt import Workbook
from openpyxl import Workbook, load_workbook
import requests
import codecs
import time
import datetime
import sys
import os
import asyncio
import json
import mysql.connector
from datetime import datetime
import pyautogui

import cv2
import numpy as np
import glob
 
base_url = "https://www.tradingview.com/chart/iSDuIFTy/"
userid = "jasarintis.com@gmail.com"
userpw = "Tmnet123!"
input_excel = "MYX.xlsx"
output_excel = "MYX_output_.xlsx"
input_data = []
screenshot_folder = "C:\\StockChartsShot"
screenshot_folder_name = "C:/StockChartsShot"

TIME_INTERVAL1 = 5
TIME_INTERVAL2 = 200

if not os.path.exists(screenshot_folder):
    os.mkdir(screenshot_folder)

today = datetime.today()
today_folder = screenshot_folder + "\\" + str(today.strftime("%Y%m%d"))
today_folder_name = screenshot_folder_name + "/" + str(today.strftime("%Y%m%d"))

if not os.path.exists(today_folder_name):
    os.mkdir(today_folder)

def input_excel_proc():
    global input_excel
    global input_data
    wb = load_workbook(input_excel)
    ws = wb.active
    for row in range(2, ws.max_row):
        exchange = ws.cell(row=row, column=1).value
        code = ws.cell(row=row, column=2).value
        stock = ws.cell(row=row, column=3).value
        bar = ws.cell(row=row, column=4).value
        data = {
            "exchange":exchange,
            "code":code,
            "stock":stock,
            "bar":bar
        }
        input_data.append(data)
            
    wb.close()

def make_video():
    img_array = []
    for filename in glob.glob( today_folder_name + '/*.png'):
        img = cv2.imread(filename)
        height, width, layers = img.shape
        size = (width,height)
        img_array.append(img)
    
    
    out = cv2.VideoWriter('screenshot_video.avi', cv2.VideoWriter_fourcc(*'DIVX'), 15, size)
    
    for i in range(len(img_array)):
        cv2.imshow('screenshot', img_array[i])
        out.write(img_array[i])
        cv2.waitKey(20)
    out.release()
    cv2.destroyAllWindows() 

input_excel_proc()

options = Options()
options.add_argument('--log-level=3')
options.add_argument("--disable-extensions")
options.add_argument("--incognito")
driver = webdriver.Chrome('chromedriver', options=options)
driver.get(base_url)

driver.implicitly_wait(3)
